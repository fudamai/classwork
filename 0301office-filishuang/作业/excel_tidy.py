#! python
# -*- coding:utf-8 -*-
# Author:fudamai
# excel_util.py


"""
按年份将Excel中数据分别保存到新的以年份命名的工作表中
"""


import os
import sys
import openpyxl
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 取给定路径上一级的路径
sys.path.append(BASE_DIR)
# 添加环境变量

file_path = os.path.join(BASE_DIR,'btc.xlsx')
# print(file_path)
wb = openpyxl.load_workbook(file_path)
sh_bt = wb['btc']


class ExcelAdmin:
    def __init__(self,sh_bt,wb,file_path):
        self.sh_bt = sh_bt
        self.wb = wb
        self.file_path = file_path
        

    def new_sheet(self):
        "迭代工作簿取年份，如果没有以年份命名的工作表，添加工作表。并将A1B1填入值"
        for row in range(2, self.sh_bt.max_row+1):  # 取行数并迭代            
            year = self.sh_bt.cell(row=row, column=1).value[0:4]
            if year not in self.wb.sheetnames:
                self.wb.create_sheet(year)
                new_sh = self.wb[year]                 
                new_sh['A1'] = self.sh_bt['A1'].value
                new_sh['B1'] = self.sh_bt['B1'].value
        # for i in self.wb.sheetnames:
            # print(i)
        return self.wb


    def data_to_new(self):
        "按年份将单元格内数据填入"        
        self.new_sheet()
        for sheet in self.wb.worksheets:
            # 迭代所有行必须在迭代工作表后进行，避免新的工作表行数与迭代的行数相同
            index = 2
            for row in self.sh_bt.rows:  # 迭代所有行
                if row[0].coordinate != 'A1' and row[0].value[0:4] == sheet.title:
                    # 匹配年份
                    print(f'{row[0]}:{row[1]}')
                    # 定义工作表类
                    sheet['A'+ str(index)] = row[0].value
                    sheet['B'+ str(index)] = row[1].value                
                    # print(f'in {sheet}:', sheet['A'+ str(index)].value, sheet['B'+ str(index)].value)
                    index += 1
        return self.wb


    def add_ave(self):
        "所有数据添加到相应的工作表后，在最后一行添加平均值"
        self.data_to_new()
        for sheet in self.wb.worksheets:
            sh = sheet
            if sh.cell(row=sh.max_row, column=1).value != '平均分':
                sh.cell(row=sh.max_row+1, column=1).value = '平均分'
                sh.cell(row=sh.max_row, column=2).value = f'=average(B2:B{sh.max_row-1})'
            print(sh.cell(row=sh.max_row, column=2))
        return self.wb


    def tidy_and_save(self):
        "将整理完的数据进行保存"
        path = self.auto_name()
        self.add_ave()
        self.wb.save(path)
        print('文件整理完毕，保存为新文件')
    
    def auto_name(self):
        "将整理过得文件保存为新文件名，新文件名在原文件名后加-1"        
        ext = self.file_path.split('.xlsx')[0]
        new_name = ext + '-1' + '.xlsx'
        return new_name


def main():
    menu ={
        'run':'启动程序',
        'quit':'关闭程序'
    }
    print('欢迎使用Excel内容处理器'.center(30, '-'))
    print('请确认将待处理的Excel文件已放入程序工作目录')
    for k,v in menu.items():
        print(f'{k}:{v}')

    while True:
        turn = input('请选择功能：').strip()
        if turn == 'run':
            ExcelAdmin(sh_bt,wb,file_path).tidy_and_save()
        elif turn == 'quit':
            break
        else:
            print('请输入正确的指令。')

if __name__ == '__main__':
    main()